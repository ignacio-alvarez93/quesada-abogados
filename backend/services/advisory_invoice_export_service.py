from __future__ import annotations

import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_DB_PATH = Path("database/quesada.db")
DEFAULT_EXPORT_DIR = Path("storage/exports/asesoria/facturas")


HEADERS = [
    "Número factura",
    "Número interno CRM",
    "Fecha factura",
    "Cliente",
    "NIF / NIE / Documento",
    "Expediente",
    "Hoja de encargo",
    "Concepto",
    "Tipo fiscal",
    "Tipo factura",
    "Factura rectificada",
    "Base imponible",
    "IVA %",
    "IVA",
    "IRPF %",
    "IRPF",
    "Suplidos",
    "Total",
    "Estado",
    "Aprobada",
    "Fecha aprobación",
    "Número cobro",
    "Fecha cobro",
    "Forma de pago",
    "Observaciones",
]


def _connect(
    db_path: Path | str = DEFAULT_DB_PATH,
) -> sqlite3.Connection:
    conn = sqlite3.connect(str(db_path), timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    conn.execute("PRAGMA busy_timeout = 30000")
    return conn


def _text(value: Any) -> str:
    return str(value or "").strip()


def _float(value: Any) -> float:
    try:
        return round(float(value or 0), 2)
    except (TypeError, ValueError):
        return 0.0


def _format_date(value: Any) -> str:
    raw = _text(value)

    if not raw:
        return ""

    for fmt in (
        "%Y-%m-%d",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%d/%m/%Y",
    ):
        try:
            return datetime.strptime(
                raw,
                fmt,
            ).strftime("%d/%m/%Y")
        except ValueError:
            continue

    return raw


def numero_factura_fiscal(
    numero_crm: Any,
) -> str:
    """
    Convierte el número interno del CRM en su representación
    visible para cliente, asesoría y libros fiscales.

    FRA-2026-0022 -> 2026-0022
    R-2026-0001   -> R-2026-0001
    """
    numero = _text(numero_crm)

    if numero.upper().startswith("FRA-"):
        return numero[4:]

    return numero


def _client_name(row: dict) -> str:
    company_name = _text(
        row.get("razon_social")
        or row.get("nombre_fiscal")
        or row.get("nombre_empresa")
        or row.get("empresa")
    )

    if company_name:
        return company_name

    return " ".join(
        part
        for part in [
            _text(row.get("nombre")),
            _text(row.get("primer_apellido")),
            _text(row.get("segundo_apellido")),
        ]
        if part
    ).strip()


def _client_document(row: dict) -> str:
    for key in (
        "nif",
        "cif",
        "nif_nie",
        "nie",
        "dni",
        "pasaporte",
        "numero_pasaporte",
        "numero_documento",
        "documento_identidad",
        "tax_id",
        "vat_number",
    ):
        value = _text(row.get(key))

        if value:
            return value.upper()

    return ""


def _tax_percentage(
    amount: Any,
    base: Any,
) -> float:
    amount_value = abs(_float(amount))
    base_value = abs(_float(base))

    if base_value <= 0:
        return 0.0

    value = amount_value * 100 / base_value

    known_values = (
        0.0,
        4.0,
        10.0,
        15.0,
        21.0,
    )

    nearest = min(
        known_values,
        key=lambda candidate: abs(candidate - value),
    )

    if abs(nearest - value) <= 0.20:
        return nearest

    return round(value, 2)


def _load_invoices(
    conn: sqlite3.Connection,
    invoice_ids: Iterable[int],
) -> list[dict]:
    resolved_ids = []

    for value in invoice_ids:
        try:
            invoice_id = int(value)
        except (TypeError, ValueError):
            continue

        if invoice_id not in resolved_ids:
            resolved_ids.append(invoice_id)

    if not resolved_ids:
        return []

    placeholders = ", ".join("?" for _ in resolved_ids)

    rows = conn.execute(
        f"""
        SELECT
            f.*,

            c.nombre,
            c.primer_apellido,
            c.segundo_apellido,
            c.nie,
            c.dni,
            c.pasaporte,

            e.numero_expediente,
            h.numero_hoja,

            fc.cobro_id,
            cob.numero_cobro,
            cob.fecha_cobro,
            cob.forma_pago,
            cob.iva_porcentaje,
            cob.irpf_porcentaje,

            original.numero_factura
                AS numero_factura_rectificada

        FROM eco_facturas f

        JOIN clientes c
          ON c.id = f.cliente_id

        LEFT JOIN expedientes e
          ON e.id = f.expediente_id

        LEFT JOIN eco_hojas_encargo h
          ON h.id = f.hoja_encargo_id

        LEFT JOIN eco_factura_cobros fc
          ON fc.id = (
              SELECT fc2.id
              FROM eco_factura_cobros fc2
              WHERE fc2.factura_id = f.id
              ORDER BY fc2.id
              LIMIT 1
          )

        LEFT JOIN eco_cobros cob
          ON cob.id = fc.cobro_id

        LEFT JOIN eco_facturas original
          ON original.id = f.factura_rectificada_id

        WHERE f.id IN ({placeholders})
          AND COALESCE(f.activo, 1) = 1

        ORDER BY
            f.fecha_factura ASC,
            f.numero_factura ASC,
            f.id ASC
        """,
        resolved_ids,
    ).fetchall()

    return [dict(row) for row in rows]


def _invoice_row(invoice: dict) -> list[Any]:
    base = _float(invoice.get("base_imponible"))
    iva = _float(invoice.get("iva"))
    irpf = _float(invoice.get("irpf"))

    iva_percentage = invoice.get("iva_porcentaje")
    irpf_percentage = invoice.get("irpf_porcentaje")

    if iva_percentage is None:
        iva_percentage = _tax_percentage(
            iva,
            base,
        )

    if irpf_percentage is None:
        irpf_percentage = _tax_percentage(
            irpf,
            base,
        )

    approved = bool(
        invoice.get("exportada_holded")
    )

    estado = _text(
        invoice.get("estado") or "BORRADOR"
    ).upper()

    if approved and estado == "EXPORTADA":
        estado = "APROBADA"

    original_number = numero_factura_fiscal(
        invoice.get("numero_factura_rectificada")
    )

    return [
        numero_factura_fiscal(
            invoice.get("numero_factura")
        ),
        _text(invoice.get("numero_factura")),
        _format_date(invoice.get("fecha_factura")),
        _client_name(invoice),
        _client_document(invoice),
        _text(invoice.get("numero_expediente")),
        _text(invoice.get("numero_hoja")),
        _text(invoice.get("concepto")),
        _text(
            invoice.get("tipo_fiscal")
            or "PROVISION"
        ).upper(),
        _text(
            invoice.get("tipo_factura")
            or "NORMAL"
        ).upper(),
        original_number,
        base,
        _float(iva_percentage),
        iva,
        _float(irpf_percentage),
        irpf,
        _float(invoice.get("suplidos")),
        _float(invoice.get("total")),
        estado,
        "Sí" if approved else "No",
        _format_date(
            invoice.get("fecha_exportacion")
        ),
        _text(invoice.get("numero_cobro")),
        _format_date(invoice.get("fecha_cobro")),
        _text(invoice.get("forma_pago")),
        _text(invoice.get("observaciones")),
    ]


def _style_sheet(worksheet) -> None:
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="0057B8",
    )
    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.row_dimensions[1].height = 38

    widths = {
        1: 18,
        2: 22,
        3: 16,
        4: 34,
        5: 22,
        6: 24,
        7: 22,
        8: 42,
        9: 16,
        10: 18,
        11: 22,
        12: 16,
        13: 12,
        14: 14,
        15: 12,
        16: 14,
        17: 14,
        18: 16,
        19: 16,
        20: 12,
        21: 20,
        22: 18,
        23: 16,
        24: 20,
        25: 42,
    }

    for column_index, width in widths.items():
        worksheet.column_dimensions[
            get_column_letter(column_index)
        ].width = width

    money_columns = (
        "L",
        "N",
        "P",
        "Q",
        "R",
    )

    percentage_columns = (
        "M",
        "O",
    )

    for column in money_columns:
        for cell in worksheet[column][1:]:
            cell.number_format = '#,##0.00 "€"'

    for column in percentage_columns:
        for cell in worksheet[column][1:]:
            cell.number_format = '0.00"%"'

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=cell.column in (4, 8, 25),
            )


def _add_summary_sheet(
    workbook: Workbook,
    invoices: list[dict],
) -> None:
    worksheet = workbook.create_sheet(
        title="Resumen"
    )

    approved_count = sum(
        1
        for invoice in invoices
        if bool(invoice.get("exportada_holded"))
    )

    base_total = round(
        sum(
            _float(invoice.get("base_imponible"))
            for invoice in invoices
        ),
        2,
    )
    iva_total = round(
        sum(
            _float(invoice.get("iva"))
            for invoice in invoices
        ),
        2,
    )
    irpf_total = round(
        sum(
            _float(invoice.get("irpf"))
            for invoice in invoices
        ),
        2,
    )
    suplidos_total = round(
        sum(
            _float(invoice.get("suplidos"))
            for invoice in invoices
        ),
        2,
    )
    total = round(
        sum(
            _float(invoice.get("total"))
            for invoice in invoices
        ),
        2,
    )

    rows = [
        ["Resumen de facturas para asesoría", ""],
        ["Generado", datetime.now().strftime("%d/%m/%Y %H:%M")],
        ["Facturas exportadas", len(invoices)],
        ["Facturas aprobadas", approved_count],
        [
            "Pendientes de aprobación",
            len(invoices) - approved_count,
        ],
        ["Base imponible", base_total],
        ["IVA", iva_total],
        ["IRPF", irpf_total],
        ["Suplidos", suplidos_total],
        ["Total facturas", total],
    ]

    for row in rows:
        worksheet.append(row)

    worksheet["A1"].font = Font(
        bold=True,
        color="FFFFFF",
    )
    worksheet["A1"].fill = PatternFill(
        fill_type="solid",
        fgColor="0057B8",
    )
    worksheet["B1"].fill = PatternFill(
        fill_type="solid",
        fgColor="0057B8",
    )

    worksheet.column_dimensions["A"].width = 30
    worksheet.column_dimensions["B"].width = 24

    for row_number in range(6, 11):
        worksheet[f"B{row_number}"].number_format = (
            '#,##0.00 "€"'
        )


def export_invoices_to_advisory(
    invoice_ids: Iterable[int],
    *,
    db_path: Path | str = DEFAULT_DB_PATH,
    output_dir: Path | str = DEFAULT_EXPORT_DIR,
) -> dict:
    """
    Exporta exclusivamente las facturas indicadas.

    No aprueba, no congela, no renumera y no modifica registros.
    """
    output_directory = Path(output_dir)
    output_directory.mkdir(
        parents=True,
        exist_ok=True,
    )

    with _connect(db_path) as conn:
        invoices = _load_invoices(
            conn,
            invoice_ids,
        )

    if not invoices:
        raise ValueError(
            "No hay facturas para exportar con los filtros actuales"
        )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Facturas"
    worksheet.append(HEADERS)

    for invoice in invoices:
        worksheet.append(
            _invoice_row(invoice)
        )

    _style_sheet(worksheet)
    _add_summary_sheet(
        workbook,
        invoices,
    )

    timestamp = datetime.now().strftime(
        "%Y%m%d_%H%M%S"
    )

    output_path = (
        output_directory
        / f"facturas_asesoria_{timestamp}.xlsx"
    )

    workbook.save(output_path)

    if (
        not output_path.exists()
        or output_path.stat().st_size <= 0
    ):
        raise RuntimeError(
            "No se pudo generar correctamente "
            "el Excel para la asesoría"
        )

    return {
        "path": str(output_path.resolve()),
        "filename": output_path.name,
        "count": len(invoices),
        "invoice_ids": [
            int(invoice["id"])
            for invoice in invoices
        ],
    }
