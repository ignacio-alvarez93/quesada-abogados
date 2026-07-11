from __future__ import annotations

import sqlite3
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_DB_PATH = Path("database/quesada.db")
DEFAULT_EXPORT_DIR = Path("storage/exports/holded")


HOLDED_HEADERS = [
    "Num factura",
    "Formato de numeración",
    "Fecha dd/mm/yyyy",
    "Fecha de vencimiento dd/mm/yyyy",
    "Fecha operación dd/mm/yyyy",
    "Descripción",
    "Nombre del contacto",
    "NIF del contacto",
    "Dirección",
    "Población",
    "Código postal",
    "Provincia",
    "País",
    "Concepto",
    "Descripción del producto",
    "SKU",
    "Precio unidad",
    "Unidades",
    "Descuento %",
    "IVA %",
    "Retención %",
    "Rec. de eq. %",
    "Operación",
    "Forma de pago (ID)",
    "Cantidad cobrada",
    "Fecha de cobro",
    "Cuenta de pago",
    "Tags separados por -",
    "Nombre canal de venta",
    "Cuenta canal de venta",
    "Moneda",
    "Cambio de moneda",
    "Almacén",
]


def _connect(db_path: Path | str = DEFAULT_DB_PATH) -> sqlite3.Connection:
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    conn.execute("PRAGMA busy_timeout = 30000")
    return conn


def _text(value: Any) -> str:
    return str(value or "").strip()


def _float(value: Any) -> float:
    try:
        return round(float(value or 0), 2)
    except (TypeError, ValueError):
        return 0.0


def _first(data: dict, *keys: str, default: Any = "") -> Any:
    for key in keys:
        value = data.get(key)

        if value is not None and str(value).strip() != "":
            return value

    return default


def _decimal(value: Any) -> Decimal:
    try:
        return Decimal(str(value or 0))
    except Exception:
        return Decimal("0")


def _holded_unit_price(
    total: Any,
    iva_pct: Any,
    irpf_pct: Any,
) -> float:
    """
    Holded recalcula impuestos desde el precio unitario.

    Se exportan seis decimales para que una factura cuyo total
    incluye IVA no pierda un céntimo al ser recalculada.
    """
    total_value = _decimal(total)
    iva_value = _decimal(iva_pct)
    irpf_value = _decimal(irpf_pct)

    divisor = (
        Decimal("1")
        + iva_value / Decimal("100")
        - irpf_value / Decimal("100")
    )

    if divisor <= 0:
        raise ValueError(
            "La combinación de IVA e IRPF no permite "
            "calcular el precio para Holded"
        )

    exact_base = total_value / divisor

    return float(
        exact_base.quantize(
            Decimal("0.000001"),
            rounding=ROUND_HALF_UP,
        )
    )


def _invoice_tax_percentages(
    invoice: dict,
) -> tuple[float, float]:
    """
    Obtiene IVA e IRPF para la exportación a Holded.

    Las facturas normales leen los porcentajes del cobro.
    Las rectificativas no tienen cobro propio, por lo que
    se deducen desde base y cuotas rectificadas.
    """
    invoice_type = _text(
        invoice.get("tipo_factura")
        or "NORMAL"
    ).upper()

    if invoice_type != "RECTIFICATIVA":
        return (
            _float(invoice.get("iva_porcentaje")),
            _float(invoice.get("irpf_porcentaje")),
        )

    base = abs(
        _float(invoice.get("base_imponible"))
    )
    iva = abs(
        _float(invoice.get("iva"))
    )
    irpf = abs(
        _float(invoice.get("irpf"))
    )

    if base <= 0:
        return 0.0, 0.0

    iva_percentage = iva * 100 / base
    irpf_percentage = irpf * 100 / base

    known_percentages = (
        0.0,
        4.0,
        10.0,
        15.0,
        21.0,
    )

    def normalize(value: float) -> float:
        nearest = min(
            known_percentages,
            key=lambda candidate: abs(
                candidate - value
            ),
        )

        # Las cuotas están guardadas a dos decimales, por lo
        # que puede aparecer 20.99 o 21.01 en lugar de 21.
        if abs(nearest - value) <= 0.20:
            return nearest

        return round(value, 4)

    return (
        normalize(iva_percentage),
        normalize(irpf_percentage),
    )


def _format_date(value: Any) -> str:
    raw = _text(value)

    if not raw:
        return ""

    for fmt in (
        "%Y-%m-%d",
        "%Y-%m-%d %H:%M:%S",
        "%d/%m/%Y",
    ):
        try:
            return datetime.strptime(raw, fmt).strftime("%d/%m/%Y")
        except ValueError:
            continue

    return raw


def _client_name(client: dict) -> str:
    company_name = _first(
        client,
        "razon_social",
        "nombre_fiscal",
        "nombre_empresa",
        "empresa",
    )

    if company_name:
        return _text(company_name)

    return " ".join(
        part
        for part in [
            _text(client.get("nombre")),
            _text(client.get("primer_apellido")),
            _text(client.get("segundo_apellido")),
        ]
        if part
    ).strip()


def _client_document(client: dict) -> str:
    """
    Identificador enviado a la columna 'NIF del contacto' de Holded.

    Prioriza documentos fiscales españoles, pero admite pasaporte
    cuando el cliente extranjero todavía no dispone de NIE.
    """
    value = _first(
        client,
        "nif",
        "cif",
        "nif_nie",
        "nie",
        "dni",
        "numero_nie",
        "numero_dni",
        "pasaporte",
        "numero_pasaporte",
        "passport",
        "passport_number",
        "numero_identificacion",
        "numero_identidad",
        "numero_documento",
        "documento_identidad",
        "documento",
        "identificacion",
        "tax_id",
        "vat_number",
    )

    return (
        _text(value)
        .upper()
        .replace(" ", "")
        .replace("-", "")
    )


def _client_nif(client: dict) -> str:
    # Alias conservado para no romper llamadas existentes.
    return _client_document(client)

def _client_address(client: dict) -> str:
    full_address = _first(
        client,
        "direccion",
        "direccion_completa",
        "domicilio",
    )

    if full_address:
        return _text(full_address)

    parts = [
        _text(
            _first(
                client,
                "tipo_via",
                "tipo_calle",
            )
        ),
        _text(
            _first(
                client,
                "nombre_via",
                "calle",
                "via",
            )
        ),
        _text(
            _first(
                client,
                "numero",
                "numero_via",
                "portal",
            )
        ),
    ]

    return " ".join(part for part in parts if part).strip()


def _load_client(conn: sqlite3.Connection, client_id: int) -> dict:
    row = conn.execute(
        """
        SELECT *
        FROM clientes
        WHERE id = ?
        """,
        (int(client_id),),
    ).fetchone()

    return dict(row) if row else {}


def _load_pending_invoices(conn: sqlite3.Connection) -> list[dict]:
    rows = conn.execute(
        """
        SELECT
            f.*,
            fc.cobro_id,
            cob.numero_cobro,
            cob.fecha_cobro,
            cob.forma_pago,
            cob.iva_porcentaje,
            cob.irpf_porcentaje,
            original.numero_factura
                AS numero_factura_rectificada
        FROM eco_facturas f
        LEFT JOIN eco_factura_cobros fc
          ON fc.id = (
              SELECT fc2.id
              FROM eco_factura_cobros fc2
              WHERE fc2.factura_id = f.id
              ORDER BY fc2.id
              LIMIT 1
          )
        LEFT JOIN eco_cobros cob
          ON cob.id = fc.cobro_id
        LEFT JOIN eco_facturas original
          ON original.id = f.factura_rectificada_id
        WHERE COALESCE(f.activo, 1) = 1
          AND COALESCE(f.exportada_holded, 0) = 0
        ORDER BY f.fecha_factura, f.numero_factura, f.id
        """
    ).fetchall()

    return [dict(row) for row in rows]


def _build_holded_row(
    invoice: dict,
    client: dict,
) -> list[Any]:
    fiscal_type = (
        "SUPLIDO"
        if _text(invoice.get("tipo_fiscal")).upper() == "SUPLIDO"
        else "PROVISION"
    )

    if fiscal_type == "SUPLIDO":
        unit_price = _float(
            invoice.get("suplidos")
            or invoice.get("total")
        )
        iva_pct, irpf_pct = _invoice_tax_percentages(
            invoice
        )
        tags = "CRM-Suplido"
    else:
        iva_pct, irpf_pct = _invoice_tax_percentages(
            invoice
        )

        unit_price = _holded_unit_price(
            invoice.get("total"),
            iva_pct,
            irpf_pct,
        )

        tags = "CRM-Provision"

    concept = (
        _text(invoice.get("concepto"))
        or "Servicios profesionales"
    )

    if (
        _text(
            invoice.get("tipo_factura")
            or "NORMAL"
        ).upper()
        == "RECTIFICATIVA"
    ):
        original_number = _text(
            invoice.get("numero_factura_rectificada")
        )

        cause = _text(
            invoice.get("causa_rectificacion")
        )

        reference_parts = [
            "Factura rectificativa",
        ]

        if original_number:
            reference_parts.append(
                f"de {original_number}"
            )

        if cause:
            reference_parts.append(cause)

        concept = " · ".join(reference_parts)

    invoice_date = _format_date(invoice.get("fecha_factura"))
    payment_date = _format_date(
        invoice.get("fecha_cobro")
        or invoice.get("fecha_factura")
    )

    invoice_type = _text(
        invoice.get("tipo_factura")
        or "NORMAL"
    ).upper()

    export_iva_pct = iva_pct
    export_irpf_pct = irpf_pct

    if invoice_type == "RECTIFICATIVA":
        # Holded necesita el IVA con signo negativo para que
        # la cuota rectificada sea negativa.
        export_iva_pct = -abs(iva_pct)

        # La retención debe mantenerse positiva. Al restarse
        # de una base negativa, produce una cuota positiva.
        export_irpf_pct = abs(irpf_pct)

    return [
        _text(invoice.get("numero_factura")),
        "",
        invoice_date,
        invoice_date,
        "",
        concept,
        _client_name(client),
        _client_nif(client),
        _client_address(client),
        _text(
            _first(
                client,
                "poblacion",
                "localidad",
                "municipio",
                "ciudad",
            )
        ),
        _text(
            _first(
                client,
                "codigo_postal",
                "cod_postal",
                "cp",
            )
        ),
        _text(
            _first(
                client,
                "provincia",
                "nombre_provincia",
            )
        ),
        _text(
            _first(
                client,
                "pais",
                "nombre_pais",
                default="España",
            )
        )
        or "España",
        concept,
        concept,
        "",
        unit_price,
        1,
        0,
        export_iva_pct,
        export_irpf_pct,
        0,
        "general",
        "",
        _float(invoice.get("total")),
        payment_date,
        "",
        tags,
        "Ventas",
        "",
        "eur",
        1,
        "",
    ]


def _style_worksheet(ws) -> None:
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="0057B8",
    )
    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 42

    widths = {
        1: 18,
        2: 20,
        3: 18,
        4: 24,
        5: 22,
        6: 34,
        7: 30,
        8: 18,
        9: 34,
        10: 20,
        11: 16,
        12: 18,
        13: 14,
        14: 34,
        15: 34,
        16: 14,
        17: 16,
        18: 12,
        19: 14,
        20: 12,
        21: 14,
        22: 14,
        23: 14,
        24: 22,
        25: 20,
        26: 18,
        27: 20,
        28: 24,
        29: 24,
        30: 24,
        31: 12,
        32: 18,
        33: 20,
    }

    for index, width in widths.items():
        ws.column_dimensions[get_column_letter(index)].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=False,
            )

    for cell in ws["Q"][1:]:
        cell.number_format = "0.000000"

    for column in ("Y", "AF"):
        for cell in ws[column][1:]:
            cell.number_format = "0.00"

    for column in ("R", "S", "T", "U", "V"):
        for cell in ws[column][1:]:
            cell.number_format = "0.00"


def export_pending_invoices_to_holded(
    *,
    db_path: Path | str = DEFAULT_DB_PATH,
    output_dir: Path | str = DEFAULT_EXPORT_DIR,
    mark_exported: bool = True,
) -> dict:
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    with _connect(db_path) as conn:
        invoices = _load_pending_invoices(conn)

        if not invoices:
            raise ValueError(
                "No hay facturas pendientes de exportar a Holded"
            )

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Holded"
        worksheet.append(HOLDED_HEADERS)

        exported_ids: list[int] = []

        for invoice in invoices:
            client = _load_client(
                conn,
                int(invoice["cliente_id"]),
            )

            worksheet.append(
                _build_holded_row(
                    invoice,
                    client,
                )
            )
            exported_ids.append(int(invoice["id"]))

        _style_worksheet(worksheet)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = (
            output_dir
            / f"facturas_emitidas_holded_{timestamp}.xlsx"
        )

        workbook.save(output_path)

        if not output_path.exists() or output_path.stat().st_size <= 0:
            raise RuntimeError(
                "No se pudo generar correctamente el archivo Holded"
            )

        if mark_exported:
            placeholders = ", ".join(
                "?"
                for _ in exported_ids
            )

            conn.execute(
                f"""
                UPDATE eco_facturas
                SET exportada_holded = 1,
                    estado = 'EXPORTADA',
                    fecha_exportacion = CURRENT_TIMESTAMP,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id IN ({placeholders})
                  AND COALESCE(exportada_holded, 0) = 0
                """,
                exported_ids,
            )
            conn.commit()

    return {
        "path": str(output_path.resolve()),
        "filename": output_path.name,
        "count": len(exported_ids),
        "invoice_ids": exported_ids,
    }
