from __future__ import annotations

from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Mapping

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_EXPORT_DIR = Path(
    "storage/exports/economico/movimientos"
)


SOURCE_LABELS = {
    "cashmatic": "Cashmatic",
    "caja_rural": "Caja Rural",
    "ing": "ING",
    "santander": "Santander",
}


HEADERS = [
    "Origen",
    "ID movimiento",
    "ID lote",
    "Número de fila",
    "ID origen",
    "Fecha operación",
    "Fecha valor / fin",
    "Concepto / motivo",
    "Referencia",
    "Operación",
    "Tipo movimiento",
    "Estado movimiento",
    "Estado revisión",
    "Importe",
    "Saldo",
    "Solicitado",
    "Introducido",
    "Dispensado",
    "No dispensado",
    "Moneda",
    "Cliente vinculado",
    "Expediente vinculado",
    "Cobro vinculado",
    "Gasto vinculado",
    "Destino vinculación",
    "Importe vinculado",
    "Pendiente vincular",
    "Estado facturación",
    "Motivo no facturable",
    "Fecha vinculación",
    "Notas",
    "Ignorado",
    "Motivo ignorado",
]


def _value(
    item: Any,
    key: str,
    default: Any = None,
) -> Any:
    if isinstance(item, Mapping):
        return item.get(key, default)

    try:
        return item[key]
    except Exception:
        return getattr(item, key, default)


def _text(value: Any) -> str:
    return str(value or "").strip()


def _integer(value: Any) -> int:
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def _euros_from_centimos(value: Any) -> float:
    return round(
        _integer(value) / 100,
        2,
    )


def _format_date(value: Any) -> str:
    raw = _text(value)

    if not raw:
        return ""

    for fmt in (
        "%Y-%m-%d",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%d/%m/%Y",
        "%d/%m/%Y %H:%M:%S",
    ):
        try:
            parsed = datetime.strptime(raw, fmt)

            if (
                parsed.hour
                or parsed.minute
                or parsed.second
            ):
                return parsed.strftime(
                    "%d/%m/%Y %H:%M:%S"
                )

            return parsed.strftime("%d/%m/%Y")

        except ValueError:
            continue

    return raw


def _source_label(source: str) -> str:
    normalized = _text(source).lower()

    return SOURCE_LABELS.get(
        normalized,
        normalized.replace("_", " ").title()
        or "Desconocido",
    )


def _movement_amount_centimos(
    source: str,
    item: Any,
) -> int:
    if source == "cashmatic":
        return _integer(
            _value(item, "net_amount_centimos")
        )

    return _integer(
        _value(item, "amount_centimos")
    )


def _movement_date(
    source: str,
    item: Any,
) -> str:
    if source == "cashmatic":
        return _format_date(
            _value(item, "start_time")
        )

    return _format_date(
        _value(item, "operation_date")
    )


def _movement_end_date(
    source: str,
    item: Any,
) -> str:
    if source == "cashmatic":
        return _format_date(
            _value(item, "end_time")
        )

    return _format_date(
        _value(item, "value_date")
    )


def _movement_concept(
    source: str,
    item: Any,
) -> str:
    if source == "cashmatic":
        return _text(
            _value(item, "reason_raw")
        )

    return _text(
        _value(item, "concept")
    )


def _movement_operation(
    source: str,
    item: Any,
) -> str:
    if source == "cashmatic":
        return _text(
            _value(item, "operation")
        )

    return ""


def _movement_type(
    source: str,
    item: Any,
) -> str:
    if source == "cashmatic":
        return _text(
            _value(item, "end_type")
        )

    return _text(
        _value(item, "movement_type")
    )


def _linked_amount_centimos(item: Any) -> int:
    return max(
        0,
        _integer(
            _value(
                item,
                "linked_amount_centimos",
            )
        ),
    )


def _pending_link_centimos(
    source: str,
    item: Any,
) -> int:
    amount = abs(
        _movement_amount_centimos(
            source,
            item,
        )
    )

    linked = _linked_amount_centimos(item)

    return max(
        0,
        amount - linked,
    )


def _movement_row(
    source: str,
    item: Any,
) -> list[Any]:
    ignored_at = _text(
        _value(item, "ignored_at")
    )

    return [
        _source_label(source),
        _value(item, "id"),
        _value(item, "batch_id"),
        _value(item, "row_number"),
        (
            _value(item, "cashmatic_id")
            if source == "cashmatic"
            else _value(item, "row_hash")
        ),
        _movement_date(source, item),
        _movement_end_date(source, item),
        _movement_concept(source, item),
        _text(
            _value(item, "reference_raw")
        ),
        _movement_operation(source, item),
        _movement_type(source, item),
        _text(
            _value(item, "movement_status")
        ),
        _text(
            _value(item, "review_status")
        ),
        _euros_from_centimos(
            _movement_amount_centimos(
                source,
                item,
            )
        ),
        (
            _euros_from_centimos(
                _value(
                    item,
                    "balance_centimos",
                )
            )
            if source != "cashmatic"
            else None
        ),
        (
            _euros_from_centimos(
                _value(
                    item,
                    "requested_centimos",
                )
            )
            if source == "cashmatic"
            else None
        ),
        (
            _euros_from_centimos(
                _value(
                    item,
                    "inserted_centimos",
                )
            )
            if source == "cashmatic"
            else None
        ),
        (
            _euros_from_centimos(
                _value(
                    item,
                    "dispensed_centimos",
                )
            )
            if source == "cashmatic"
            else None
        ),
        (
            _euros_from_centimos(
                _value(
                    item,
                    "not_dispensed_centimos",
                )
            )
            if source == "cashmatic"
            else None
        ),
        _text(
            _value(item, "currency") or "EUR"
        ),
        _value(item, "linked_client_id"),
        _value(item, "linked_expedient_id"),
        _value(item, "linked_payment_id"),
        _value(item, "linked_gasto_id"),
        _text(
            _value(item, "linked_target_type")
        ),
        _euros_from_centimos(
            _linked_amount_centimos(item)
        ),
        _euros_from_centimos(
            _pending_link_centimos(
                source,
                item,
            )
        ),
        _text(
            _value(
                item,
                "invoiceability_status",
            )
        ),
        _text(
            _value(
                item,
                "invoiceability_reason",
            )
        ),
        _format_date(
            _value(item, "linked_at")
        ),
        _text(
            _value(item, "link_notes")
        ),
        "Sí" if ignored_at else "No",
        _text(
            _value(item, "ignored_reason")
        ),
    ]


def _style_movements_sheet(worksheet) -> None:
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="0057B8",
    )
    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.row_dimensions[1].height = 38

    widths = {
        1: 18,
        2: 15,
        3: 12,
        4: 12,
        5: 34,
        6: 22,
        7: 22,
        8: 55,
        9: 30,
        10: 20,
        11: 28,
        12: 34,
        13: 28,
        14: 16,
        15: 16,
        16: 16,
        17: 16,
        18: 16,
        19: 16,
        20: 12,
        21: 18,
        22: 20,
        23: 18,
        24: 18,
        25: 22,
        26: 18,
        27: 18,
        28: 22,
        29: 40,
        30: 22,
        31: 45,
        32: 12,
        33: 40,
    }

    for index, width in widths.items():
        worksheet.column_dimensions[
            get_column_letter(index)
        ].width = width

    money_columns = (
        "N",
        "O",
        "P",
        "Q",
        "R",
        "S",
        "Z",
        "AA",
    )

    for column in money_columns:
        for cell in worksheet[column][1:]:
            cell.number_format = '#,##0.00 "€"'

    wrap_columns = {
        5,
        8,
        9,
        12,
        13,
        28,
        29,
        31,
        33,
    }

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=(
                    cell.column in wrap_columns
                ),
            )


def _add_summary_sheet(
    workbook: Workbook,
    *,
    source: str,
    movements: list[Any],
    search: str,
) -> None:
    worksheet = workbook.create_sheet(
        title="Resumen"
    )

    amounts = [
        _movement_amount_centimos(
            source,
            movement,
        )
        for movement in movements
    ]

    income_centimos = sum(
        amount
        for amount in amounts
        if amount > 0
    )

    expense_centimos = sum(
        abs(amount)
        for amount in amounts
        if amount < 0
    )

    net_centimos = sum(amounts)

    linked_count = sum(
        1
        for movement in movements
        if (
            _value(
                movement,
                "linked_payment_id",
            )
            or _value(
                movement,
                "linked_gasto_id",
            )
            or _linked_amount_centimos(
                movement
            )
            > 0
        )
    )

    ignored_count = sum(
        1
        for movement in movements
        if _text(
            _value(movement, "ignored_at")
        )
    )

    status_counts = Counter(
        _text(
            _value(
                movement,
                "movement_status",
            )
        )
        or "SIN_ESTADO"
        for movement in movements
    )

    rows = [
        ["Resumen de movimientos", ""],
        [
            "Generado",
            datetime.now().strftime(
                "%d/%m/%Y %H:%M"
            ),
        ],
        ["Origen", _source_label(source)],
        [
            "Filtro de búsqueda",
            search or "Sin filtro",
        ],
        ["Movimientos exportados", len(movements)],
        ["Ingresos", income_centimos / 100],
        ["Salidas", expense_centimos / 100],
        ["Saldo neto", net_centimos / 100],
        ["Con vinculación", linked_count],
        [
            "Sin vinculación",
            len(movements) - linked_count,
        ],
        ["Ignorados", ignored_count],
    ]

    for row in rows:
        worksheet.append(row)

    start_status_row = len(rows) + 3

    worksheet.cell(
        row=start_status_row,
        column=1,
        value="Desglose por estado",
    )
    worksheet.cell(
        row=start_status_row,
        column=2,
        value="Movimientos",
    )

    for offset, (
        status,
        count,
    ) in enumerate(
        sorted(status_counts.items()),
        start=1,
    ):
        worksheet.cell(
            row=start_status_row + offset,
            column=1,
            value=status,
        )
        worksheet.cell(
            row=start_status_row + offset,
            column=2,
            value=count,
        )

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="0057B8",
    )
    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for cell in (
        worksheet["A1"],
        worksheet["B1"],
        worksheet.cell(
            row=start_status_row,
            column=1,
        ),
        worksheet.cell(
            row=start_status_row,
            column=2,
        ),
    ):
        cell.fill = header_fill
        cell.font = header_font

    worksheet.column_dimensions["A"].width = 34
    worksheet.column_dimensions["B"].width = 34

    for row_number in range(6, 9):
        worksheet[f"B{row_number}"].number_format = (
            '#,##0.00 "€"'
        )


def export_movements_to_excel(
    source: str,
    movements: Iterable[Any],
    *,
    search: str = "",
    output_dir: Path | str = DEFAULT_EXPORT_DIR,
) -> dict[str, Any]:
    normalized_source = _text(source).lower()

    if normalized_source not in SOURCE_LABELS:
        raise ValueError(
            "El origen de movimientos no es válido"
        )

    movement_list = list(movements)

    if not movement_list:
        raise ValueError(
            "No hay movimientos para exportar "
            "con los filtros actuales"
        )

    output_directory = Path(output_dir)
    output_directory.mkdir(
        parents=True,
        exist_ok=True,
    )

    timestamp = datetime.now().strftime(
        "%Y%m%d_%H%M%S"
    )

    output_path = output_directory / (
        f"movimientos_{normalized_source}_"
        f"{timestamp}.xlsx"
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Movimientos"

    worksheet.append(HEADERS)

    for movement in movement_list:
        worksheet.append(
            _movement_row(
                normalized_source,
                movement,
            )
        )

    _style_movements_sheet(worksheet)

    _add_summary_sheet(
        workbook,
        source=normalized_source,
        movements=movement_list,
        search=_text(search),
    )

    workbook.save(output_path)

    return {
        "path": str(output_path.resolve()),
        "count": len(movement_list),
        "source": normalized_source,
        "source_label": _source_label(
            normalized_source
        ),
    }
