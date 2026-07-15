from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Mapping

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_EXPORT_DIR = Path(
    "storage/exports/economico/gastos"
)


HEADERS = [
    "ID",
    "Fecha gasto",
    "Fecha factura",
    "Proveedor",
    "NIF/CIF",
    "Número factura",
    "Concepto",
    "Categoría",
    "Tipo justificante",
    "Forma de pago",
    "Base imponible",
    "IVA %",
    "IVA",
    "IRPF %",
    "IRPF",
    "Otros impuestos / ajustes",
    "Total",
    "IVA deducible",
    "Deducible IRPF",
    "Porcentaje deducible",
    "Estado fiscal",
    "Estado documental",
    "Estado conciliación",
    "Expediente",
    "Movimiento bancario",
    "Ruta documento",
    "Observaciones",
]


def _value(
    item: Any,
    key: str,
    default: Any = None,
) -> Any:
    if isinstance(item, Mapping):
        return item.get(key, default)

    try:
        return item[key]
    except Exception:
        return getattr(item, key, default)


def _text(value: Any) -> str:
    return str(value or "").strip()


def _integer(value: Any) -> int:
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def _float(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _euros(value_centimos: Any) -> float:
    return round(
        _integer(value_centimos) / 100,
        2,
    )


def _format_date(value: Any) -> str:
    raw = _text(value)

    if not raw:
        return ""

    for fmt in (
        "%Y-%m-%d",
        "%Y-%m-%d %H:%M:%S",
        "%d/%m/%Y",
    ):
        try:
            return datetime.strptime(
                raw,
                fmt,
            ).strftime("%d/%m/%Y")
        except ValueError:
            continue

    return raw


def _yes_no(value: Any) -> str:
    return "Sí" if bool(value) else "No"


def _expense_row(
    expense: Any,
) -> list[Any]:
    total_centimos = _integer(
        _value(
            expense,
            "effective_total_centimos",
            _value(expense, "total_centimos"),
        )
    )

    return [
        _value(expense, "id"),
        _format_date(
            _value(expense, "fecha_gasto")
        ),
        _format_date(
            _value(expense, "fecha_factura")
        ),
        _text(
            _value(
                expense,
                "supplier_display_name",
                _value(
                    expense,
                    "supplier_name_snapshot",
                    _value(expense, "proveedor"),
                ),
            )
        ),
        _text(
            _value(
                expense,
                "supplier_tax_id",
                _value(
                    expense,
                    "supplier_tax_id_snapshot",
                ),
            )
        ),
        _text(
            _value(expense, "numero_factura")
        ),
        _text(
            _value(expense, "concepto")
        ),
        _text(
            _value(expense, "categoria")
        ),
        _text(
            _value(expense, "tipo_justificante")
        ),
        _text(
            _value(expense, "forma_pago")
        ),
        _euros(
            _value(
                expense,
                "base_imponible_centimos",
            )
        ),
        _float(
            _value(expense, "iva_porcentaje")
        ),
        _euros(
            _value(expense, "iva_centimos")
        ),
        _float(
            _value(expense, "irpf_porcentaje")
        ),
        _euros(
            _value(expense, "irpf_centimos")
        ),
        _euros(
            _value(
                expense,
                "otros_impuestos_centimos",
            )
        ),
        _euros(total_centimos),
        _yes_no(
            _value(expense, "iva_deducible")
        ),
        _yes_no(
            _value(
                expense,
                "deducible_irpf",
                _value(expense, "deducible"),
            )
        ),
        _float(
            _value(
                expense,
                "porcentaje_deducible",
            )
        ),
        _text(
            _value(expense, "estado_fiscal")
        ),
        _text(
            _value(
                expense,
                "estado_documental",
            )
        ),
        _text(
            _value(
                expense,
                "estado_conciliacion",
            )
        ),
        _text(
            _value(expense, "numero_expediente")
        ),
        _value(expense, "bank_movement_id"),
        _text(
            _value(
                expense,
                "documento_ruta",
                _value(
                    expense,
                    "factura_recibida_ruta",
                ),
            )
        ),
        _text(
            _value(expense, "observaciones")
        ),
    ]


def _style_expenses_sheet(
    worksheet,
) -> None:
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="0057B8",
    )
    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.row_dimensions[1].height = 36

    widths = {
        1: 10,
        2: 16,
        3: 16,
        4: 34,
        5: 18,
        6: 20,
        7: 45,
        8: 30,
        9: 22,
        10: 20,
        11: 17,
        12: 12,
        13: 17,
        14: 12,
        15: 17,
        16: 20,
        17: 17,
        18: 16,
        19: 18,
        20: 18,
        21: 24,
        22: 26,
        23: 26,
        24: 22,
        25: 20,
        26: 50,
        27: 55,
    }

    for index, width in widths.items():
        worksheet.column_dimensions[
            get_column_letter(index)
        ].width = width

    money_columns = (
        "K",
        "M",
        "O",
        "P",
        "Q",
    )

    for column in money_columns:
        for cell in worksheet[column][1:]:
            cell.number_format = '#,##0.00 "€"'

    percentage_columns = (
        "L",
        "N",
        "T",
    )

    for column in percentage_columns:
        for cell in worksheet[column][1:]:
            cell.number_format = '0.00'

    wrap_columns = {
        4,
        7,
        8,
        21,
        22,
        23,
        26,
        27,
    }

    for row in worksheet.iter_rows(
        min_row=2,
    ):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=(
                    cell.column in wrap_columns
                ),
            )


def _add_summary_sheet(
    workbook: Workbook,
    expenses: list[Any],
    *,
    search: str,
    quick_filter: str,
    date_from: str | None,
    date_to: str | None,
) -> None:
    worksheet = workbook.create_sheet(
        title="Resumen"
    )

    total_centimos = sum(
        _integer(
            _value(
                expense,
                "effective_total_centimos",
                _value(
                    expense,
                    "total_centimos",
                ),
            )
        )
        for expense in expenses
    )

    base_centimos = sum(
        _integer(
            _value(
                expense,
                "base_imponible_centimos",
            )
        )
        for expense in expenses
    )

    iva_centimos = sum(
        _integer(
            _value(expense, "iva_centimos")
        )
        for expense in expenses
    )

    irpf_centimos = sum(
        _integer(
            _value(expense, "irpf_centimos")
        )
        for expense in expenses
    )

    pending_centimos = sum(
        _integer(
            _value(
                expense,
                "effective_total_centimos",
                _value(
                    expense,
                    "total_centimos",
                ),
            )
        )
        for expense in expenses
        if _text(
            _value(
                expense,
                "estado_conciliacion",
            )
        ).upper()
        in {"PENDIENTE", "PARCIAL"}
    )

    rows = [
        ["Resumen de gastos", ""],
        [
            "Generado",
            datetime.now().strftime(
                "%d/%m/%Y %H:%M"
            ),
        ],
        [
            "Búsqueda",
            search or "Sin búsqueda",
        ],
        [
            "Filtro rápido",
            quick_filter or "ALL",
        ],
        [
            "Desde",
            _format_date(date_from),
        ],
        [
            "Hasta",
            _format_date(date_to),
        ],
        [
            "Gastos exportados",
            len(expenses),
        ],
        [
            "Base imponible",
            base_centimos / 100,
        ],
        [
            "IVA soportado",
            iva_centimos / 100,
        ],
        [
            "IRPF",
            irpf_centimos / 100,
        ],
        [
            "Total gastos",
            total_centimos / 100,
        ],
        [
            "Pendiente de conciliar",
            pending_centimos / 100,
        ],
    ]

    for row in rows:
        worksheet.append(row)

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="0057B8",
    )
    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    worksheet.column_dimensions["A"].width = 30
    worksheet.column_dimensions["B"].width = 36

    for row_number in range(8, 13):
        worksheet[f"B{row_number}"].number_format = (
            '#,##0.00 "€"'
        )


def export_expenses_to_excel(
    expenses: Iterable[Any],
    *,
    search: str = "",
    quick_filter: str = "ALL",
    date_from: str | None = None,
    date_to: str | None = None,
    output_dir: Path | str = DEFAULT_EXPORT_DIR,
) -> dict[str, Any]:
    expense_list = list(expenses)

    if not expense_list:
        raise ValueError(
            "No hay gastos para exportar "
            "con los filtros actuales."
        )

    output_directory = Path(output_dir)
    output_directory.mkdir(
        parents=True,
        exist_ok=True,
    )

    timestamp = datetime.now().strftime(
        "%Y%m%d_%H%M%S"
    )

    output_path = output_directory / (
        f"gastos_{timestamp}.xlsx"
    )

    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = "Gastos"
    worksheet.append(HEADERS)

    for expense in expense_list:
        worksheet.append(
            _expense_row(expense)
        )

    _style_expenses_sheet(worksheet)

    _add_summary_sheet(
        workbook,
        expense_list,
        search=search,
        quick_filter=quick_filter,
        date_from=date_from,
        date_to=date_to,
    )

    workbook.save(output_path)

    return {
        "path": str(output_path.resolve()),
        "count": len(expense_list),
        "filename": output_path.name,
    }
