from __future__ import annotations

import csv
import hashlib
import json
from dataclasses import dataclass, asdict
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Iterable


EXPECTED_COLUMNS = [
    "ID",
    "REQUESTED",
    "INSERTED",
    "DISPENSED",
    "NOT_DISPENSED",
    "CURRENCY",
    "OPERATION",
    "RESULT",
    "END_TYPE",
    "MODULE_TYPE",
    "START_TIME",
    "END_TIME",
    "SOURCE",
    "REASON",
    "REFERENCE",
    "USER_ID",
    "USER_USERNAME",
    "USER_FIRST_NAME",
    "USER_LAST_NAME",
    "USER_EMAIL",
]

ALLOWED_OPERATIONS = {
    # Operaciones de cobro
    "payment",

    # Operaciones internas Cashmatic
    "cashbox",
    "cashbox_notes",
    "register_closure",
    "refill",
    "withdrawal",
    "withdrawal_denominations",
    "empty",
    "empty_notes",
}


@dataclass(frozen=True)
class CashmaticDiagnosticRow:
    row_number: int
    row_hash: str
    cashmatic_id: str
    operation: str
    result: str
    end_type: str
    requested_centimos: int
    inserted_centimos: int
    dispensed_centimos: int
    not_dispensed_centimos: int
    net_amount_centimos: int
    currency: str
    start_time: str
    end_time: str
    reason_raw: str
    reference_raw: str
    source_raw: str
    user_username: str
    candidate_payment: bool
    status: str
    warnings: list[str]


@dataclass(frozen=True)
class CashmaticDiagnosticReport:
    file_path: str
    file_name: str
    file_sha256: str
    detected_format: str
    total_rows: int
    valid_rows: int
    quarantine_rows: int
    candidate_payment_rows: int
    operations_count: dict[str, int]
    status_count: dict[str, int]
    total_candidate_requested_centimos: int
    total_candidate_inserted_centimos: int
    total_candidate_dispensed_centimos: int
    total_candidate_net_centimos: int
    first_start_time: str | None
    last_start_time: str | None
    sample_quarantine: list[dict[str, Any]]
    manual_linking_policy: str


def file_sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def stable_row_hash(row: dict[str, Any]) -> str:
    payload = json.dumps(row, ensure_ascii=False, sort_keys=True, default=str)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def parse_money_to_centimos(value: Any) -> int:
    if value is None:
        return 0

    if isinstance(value, int):
        return value

    if isinstance(value, float):
        # Cashmatic exporta importes como enteros en céntimos.
        # Si Excel lo trae como 5000.0, lo convertimos a 5000.
        return int(round(value))

    text = str(value).strip()
    if not text:
        return 0

    text = text.replace("\u00a0", "").replace(" ", "")

    # Caso español tipo "1.234,56": convertir a euros -> céntimos.
    if "," in text and "." in text:
        normalized = text.replace(".", "").replace(",", ".")
        return int(round(float(normalized) * 100))

    # Caso "1234,56": euros -> céntimos.
    if "," in text:
        normalized = text.replace(",", ".")
        return int(round(float(normalized) * 100))

    # Caso habitual Cashmatic: "5000" = 50,00 €.
    try:
        return int(float(text))
    except ValueError:
        return 0


def excel_serial_to_datetime(value: int | float) -> datetime:
    # Excel Windows serial date. 25569 = 1970-01-01.
    return datetime(1899, 12, 30) + timedelta(days=float(value))


def parse_datetime_to_iso(value: Any) -> str:
    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")

    if isinstance(value, (int, float)):
        try:
            return excel_serial_to_datetime(value).isoformat(sep=" ", timespec="seconds")
        except Exception:
            return str(value)

    text = str(value).strip()
    if not text:
        return ""

    formats = [
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%H:%M:%S %d/%m/%Y",
        "%H:%M %d/%m/%Y",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%dT%H:%M:%S",
    ]

    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).isoformat(sep=" ", timespec="seconds")
        except ValueError:
            pass

    return text


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def validate_cashmatic_headers(headers: list[str], path: Path) -> None:
    """Valida el contrato estable de exportación Cashmatic.

    Cashmatic exporta siempre 20 columnas conocidas. Si el proveedor cambia
    el formato, preferimos fallar de forma explícita antes que importar datos
    desplazados o corruptos.
    """
    normalized_headers = [clean_text(x) for x in headers if clean_text(x)]

    if normalized_headers != EXPECTED_COLUMNS:
        missing = [col for col in EXPECTED_COLUMNS if col not in normalized_headers]
        unexpected = [col for col in normalized_headers if col not in EXPECTED_COLUMNS]
        raise ValueError(
            "Formato Cashmatic no reconocido en "
            f"{path}. Columnas esperadas={EXPECTED_COLUMNS}. "
            f"Columnas recibidas={normalized_headers}. "
            f"Faltan={missing}. Sobran={unexpected}."
        )


def read_csv_rows(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []

    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        sample = fh.read(4096)
        fh.seek(0)

        delimiter = ";"
        if sample.count(",") > sample.count(";"):
            delimiter = ","

        reader = csv.DictReader(fh, delimiter=delimiter)
        validate_cashmatic_headers(list(reader.fieldnames or []), path)

        for raw in reader:
            normalized: dict[str, Any] = {}
            for col in EXPECTED_COLUMNS:
                normalized[col] = raw.get(col, "")

            # Algunos CSV terminan con ; y DictReader crea clave None.
            rows.append(normalized)

    return rows


def read_xlsx_rows(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "Falta openpyxl para leer XLSX. Instala con: pip install openpyxl"
        ) from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    iterator = ws.iter_rows(values_only=True)
    headers_raw = next(iterator, None)
    if not headers_raw:
        return []

    headers = [clean_text(x) for x in headers_raw]
    validate_cashmatic_headers(headers, path)

    rows: list[dict[str, Any]] = []

    for values in iterator:
        raw = dict(zip(headers, values))
        normalized: dict[str, Any] = {}
        for col in EXPECTED_COLUMNS:
            normalized[col] = raw.get(col, "")
        rows.append(normalized)

    return rows


def read_cashmatic_rows(path: Path) -> tuple[str, list[dict[str, Any]]]:
    suffix = path.suffix.lower()

    if suffix == ".csv":
        return "csv", read_csv_rows(path)

    if suffix in {".xlsx", ".xlsm"}:
        return "xlsx", read_xlsx_rows(path)

    raise ValueError(f"Formato no soportado: {suffix}. Usa CSV o XLSX.")


def normalize_cashmatic_row(row_number: int, raw: dict[str, Any]) -> CashmaticDiagnosticRow:
    operation = clean_text(raw.get("OPERATION")).lower()
    result = clean_text(raw.get("RESULT"))
    end_type = clean_text(raw.get("END_TYPE")).lower()

    requested = parse_money_to_centimos(raw.get("REQUESTED"))
    inserted = parse_money_to_centimos(raw.get("INSERTED"))
    dispensed = parse_money_to_centimos(raw.get("DISPENSED"))
    not_dispensed = parse_money_to_centimos(raw.get("NOT_DISPENSED"))
    net_amount = inserted - dispensed

    warnings: list[str] = []

    if operation not in ALLOWED_OPERATIONS:
        warnings.append(f"Operación no permitida o desconocida: {operation or '-'}")

    if operation == "payment" and result != "1":
        warnings.append("Pago con RESULT distinto de 1")

    if operation == "payment" and end_type != "normal":
        warnings.append(f"Pago con END_TYPE no normal: {end_type or '-'}")

    if operation == "payment" and net_amount <= 0:
        warnings.append("Pago con importe neto no positivo")

    if operation == "payment" and result == "1" and end_type == "normal" and requested != net_amount:
        warnings.append("REQUESTED no coincide con INSERTED - DISPENSED")

    candidate_payment = (
        operation == "payment"
        and result == "1"
        and end_type == "normal"
        and net_amount > 0
    )

    if operation not in ALLOWED_OPERATIONS:
        status = "QUARANTINE"
    elif candidate_payment:
        status = "CANDIDATE_PAYMENT_MANUAL_LINK_REQUIRED"
    elif operation == "payment":
        status = "PAYMENT_REVIEW_REQUIRED"
    else:
        status = "INTERNAL_CASHMATIC_MOVEMENT"

    return CashmaticDiagnosticRow(
        row_number=row_number,
        row_hash=stable_row_hash(raw),
        cashmatic_id=clean_text(raw.get("ID")),
        operation=operation,
        result=result,
        end_type=end_type,
        requested_centimos=requested,
        inserted_centimos=inserted,
        dispensed_centimos=dispensed,
        not_dispensed_centimos=not_dispensed,
        net_amount_centimos=net_amount,
        currency=clean_text(raw.get("CURRENCY")),
        start_time=parse_datetime_to_iso(raw.get("START_TIME")),
        end_time=parse_datetime_to_iso(raw.get("END_TIME")),
        reason_raw=clean_text(raw.get("REASON")),
        reference_raw=clean_text(raw.get("REFERENCE")),
        source_raw=clean_text(raw.get("SOURCE")),
        user_username=clean_text(raw.get("USER_USERNAME")),
        candidate_payment=candidate_payment,
        status=status,
        warnings=warnings,
    )


def cents_to_eur(value: int) -> str:
    sign = "-" if value < 0 else ""
    value = abs(value)
    euros = value // 100
    cents = value % 100
    return f"{sign}{euros}.{cents:02d}"


def diagnose_cashmatic_file(path: str | Path) -> tuple[CashmaticDiagnosticReport, list[CashmaticDiagnosticRow]]:
    file_path = Path(path)

    if not file_path.exists():
        raise FileNotFoundError(f"No existe el archivo: {file_path}")

    detected_format, raw_rows = read_cashmatic_rows(file_path)
    normalized_rows = [
        normalize_cashmatic_row(index + 2, row)
        for index, row in enumerate(raw_rows)
    ]

    operations_count: dict[str, int] = {}
    status_count: dict[str, int] = {}

    for row in normalized_rows:
        operations_count[row.operation or "-"] = operations_count.get(row.operation or "-", 0) + 1
        status_count[row.status] = status_count.get(row.status, 0) + 1

    candidates = [row for row in normalized_rows if row.candidate_payment]
    quarantine = [row for row in normalized_rows if row.status == "QUARANTINE"]

    starts = sorted([row.start_time for row in normalized_rows if row.start_time])

    report = CashmaticDiagnosticReport(
        file_path=str(file_path),
        file_name=file_path.name,
        file_sha256=file_sha256(file_path),
        detected_format=detected_format,
        total_rows=len(normalized_rows),
        valid_rows=len([row for row in normalized_rows if row.status != "QUARANTINE"]),
        quarantine_rows=len(quarantine),
        candidate_payment_rows=len(candidates),
        operations_count=dict(sorted(operations_count.items())),
        status_count=dict(sorted(status_count.items())),
        total_candidate_requested_centimos=sum(row.requested_centimos for row in candidates),
        total_candidate_inserted_centimos=sum(row.inserted_centimos for row in candidates),
        total_candidate_dispensed_centimos=sum(row.dispensed_centimos for row in candidates),
        total_candidate_net_centimos=sum(row.net_amount_centimos for row in candidates),
        first_start_time=starts[0] if starts else None,
        last_start_time=starts[-1] if starts else None,
        sample_quarantine=[asdict(row) for row in quarantine[:10]],
        manual_linking_policy=(
            "La vinculación con cliente, expediente o cobro NO es automática. "
            "El sistema solo marca candidatos y conserva REASON/REFERENCE como pistas para revisión manual."
        ),
    )

    return report, normalized_rows


def report_to_dict(report: CashmaticDiagnosticReport) -> dict[str, Any]:
    data = asdict(report)
    data["total_candidate_requested_eur"] = cents_to_eur(report.total_candidate_requested_centimos)
    data["total_candidate_inserted_eur"] = cents_to_eur(report.total_candidate_inserted_centimos)
    data["total_candidate_dispensed_eur"] = cents_to_eur(report.total_candidate_dispensed_centimos)
    data["total_candidate_net_eur"] = cents_to_eur(report.total_candidate_net_centimos)
    return data
