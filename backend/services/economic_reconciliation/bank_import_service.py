from __future__ import annotations

import sqlite3
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from backend.services.economic_reconciliation.bank_santander_parser_service import (
    SantanderBankDiagnosticReport,
    SantanderBankDiagnosticRow,
    diagnose_santander_bank_file,
)
from backend.services.economic_reconciliation.bank_caja_rural_parser_service import (
    CajaRuralBankDiagnosticReport,
    CajaRuralBankDiagnosticRow,
    diagnose_caja_rural_bank_file,
)
from backend.services.economic_reconciliation.bank_ing_parser_service import (
    IngBankDiagnosticReport,
    IngBankDiagnosticRow,
    diagnose_ing_bank_file,
)
from backend.services.economic_reconciliation.cashmatic_import_service import (
    DEFAULT_DB_PATH,
    connect,
    ensure_schema as ensure_cashmatic_schema,
)


BANK_MIGRATION_PATH = Path("database/migrations/20260706_create_bank_reconciliation_staging.sql")


@dataclass(frozen=True)
class BankImportResult:
    batch_id: int
    batch_created: bool
    source_file: str
    file_sha256: str
    total_rows: int
    inserted_rows: int
    duplicate_rows: int
    income_rows: int
    expense_rows: int
    quarantine_rows: int
    manual_linking_policy: str


def ensure_bank_schema(conn: sqlite3.Connection) -> None:
    # economic_import_batches vive en la migración Cashmatic; la reutilizamos.
    ensure_cashmatic_schema(conn)

    if not BANK_MIGRATION_PATH.exists():
        raise FileNotFoundError(f"No existe migración bancaria: {BANK_MIGRATION_PATH}")

    conn.executescript(BANK_MIGRATION_PATH.read_text(encoding="utf-8"))
    conn.commit()



def _file_sha256_flexible(path):
    import hashlib
    from pathlib import Path

    h = hashlib.sha256()
    with Path(path).open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def _normalize_bank_header_flexible(value: object) -> str:
    import unicodedata

    raw = str(value or "").strip().lower()
    raw = unicodedata.normalize("NFKD", raw)
    raw = "".join(ch for ch in raw if not unicodedata.combining(ch))
    raw = raw.replace("\n", " ").replace("\r", " ")
    raw = raw.replace(".", "")
    raw = " ".join(raw.split())
    return raw


def _date_flexible_to_sql(value: object) -> str:
    from datetime import datetime, date, timedelta

    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")

    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")

    raw = str(value or "").strip()
    if not raw:
        return ""

    try:
        if raw.replace(".", "", 1).isdigit():
            serial = float(raw)
            if serial > 1000:
                return (datetime(1899, 12, 30) + timedelta(days=serial)).strftime("%Y-%m-%d")
    except Exception:
        pass

    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(raw[:10], fmt).strftime("%Y-%m-%d")
        except Exception:
            pass

    return raw[:10]


def _amount_flexible_to_centimos(value: object) -> int:
    from decimal import Decimal, ROUND_HALF_UP

    if value is None:
        return 0

    raw = str(value).strip()
    if not raw:
        return 0

    raw = raw.replace("€", "").replace(" ", "")

    if "," in raw and "." in raw:
        raw = raw.replace(".", "").replace(",", ".")
    elif "," in raw:
        raw = raw.replace(",", ".")

    return int((Decimal(raw) * Decimal("100")).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def _sha256_bank_payload(value: str) -> str:
    import hashlib
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def _first_matching_column(header_map: dict[str, int], candidates: list[str]) -> int | None:
    normalized_candidates = [_normalize_bank_header_flexible(c) for c in candidates]
    for candidate in normalized_candidates:
        if candidate in header_map:
            return header_map[candidate]
    return None


def _find_caja_rural_flexible_header(raw_rows: list[tuple]) -> tuple[int, dict[str, int]] | None:
    for idx, row in enumerate(raw_rows[:25]):
        header_map = {_normalize_bank_header_flexible(value): col_idx for col_idx, value in enumerate(row)}

        date_idx = _first_matching_column(
            header_map,
            ["Fecha de la operación", "Fecha operacion", "Fecha Ejecución", "Fecha ejecucion"],
        )
        value_date_idx = _first_matching_column(
            header_map,
            ["Fecha valor", "Fecha Valor"],
        )
        concept_idx = _first_matching_column(
            header_map,
            ["Tipo movimiento", "Descripcion", "Descripción", "Concepto"],
        )
        amount_idx = _first_matching_column(
            header_map,
            ["Importe"],
        )
        balance_idx = _first_matching_column(
            header_map,
            ["Saldo"],
        )

        if date_idx is not None and value_date_idx is not None and concept_idx is not None and amount_idx is not None and balance_idx is not None:
            statement_idx = _first_matching_column(
                header_map,
                ["Nro. Apunte", "Nro Apunte", "Número apunte", "Numero apunte", "Apunte"],
            )
            return idx, {
                "operation_date": date_idx,
                "value_date": value_date_idx,
                "concept": concept_idx,
                "amount": amount_idx,
                "balance": balance_idx,
                "statement_number": statement_idx if statement_idx is not None else -1,
            }

    return None


def diagnose_caja_rural_bank_file_flexible(path):
    from pathlib import Path
    from openpyxl import load_workbook

    source_file_path = Path(path)
    workbook = load_workbook(source_file_path, data_only=True)
    worksheet = workbook.active

    raw_rows = list(worksheet.iter_rows(values_only=True))
    header_info = _find_caja_rural_flexible_header(raw_rows)

    if header_info is None:
        raise ValueError(
            "Formato Caja Rural no reconocido. No se encontró cabecera compatible "
            "en las primeras 25 filas."
        )

    header_idx, columns = header_info

    parsed_rows = []
    quarantine_rows = 0
    income_rows = 0
    expense_rows = 0
    total_income_centimos = 0
    total_expense_centimos = 0
    by_type = {}
    by_status = {}

    for row_number, row in enumerate(raw_rows[header_idx + 1:], start=header_idx + 2):
        try:
            operation_date = _date_flexible_to_sql(row[columns["operation_date"]])
            value_date = _date_flexible_to_sql(row[columns["value_date"]])
            concept = str(row[columns["concept"]] or "").strip()
            amount_centimos = _amount_flexible_to_centimos(row[columns["amount"]])
            balance_centimos = _amount_flexible_to_centimos(row[columns["balance"]])

            statement_number = ""
            statement_idx = columns.get("statement_number", -1)
            if statement_idx is not None and statement_idx >= 0 and statement_idx < len(row):
                raw_statement = row[statement_idx]
                statement_number = str(raw_statement or "").strip()
                if statement_number.endswith(".0"):
                    statement_number = statement_number[:-2]

            if not operation_date or not concept:
                quarantine_rows += 1
                continue

            if amount_centimos > 0:
                movement_type = "INCOME"
                movement_status = "BANK_INCOME_REVIEW_REQUIRED"
                income_rows += 1
                total_income_centimos += amount_centimos
            elif amount_centimos < 0:
                movement_type = "EXPENSE"
                movement_status = "BANK_EXPENSE_REVIEW_REQUIRED"
                expense_rows += 1
                total_expense_centimos += amount_centimos
            else:
                movement_type = "ZERO"
                movement_status = "BANK_ZERO_REVIEW_REQUIRED"

            by_type[movement_type] = by_type.get(movement_type, 0) + 1
            by_status[movement_status] = by_status.get(movement_status, 0) + 1

            # Si hay Nro. Apunte, lo usamos como parte fuerte del hash.
            # Si no hay, usamos saldo para distinguir apuntes repetidos.
            hash_parts = [
                "CAJA_RURAL_FLEX",
                operation_date,
                value_date,
                concept.lower(),
                str(amount_centimos),
                str(balance_centimos),
                str(statement_number),
            ]
            row_hash = _sha256_bank_payload("|".join(hash_parts))

            parsed_rows.append(
                CajaRuralBankDiagnosticRow(
                    row_number=row_number,
                    operation_date=operation_date,
                    value_date=value_date,
                    concept=concept,
                    amount_centimos=amount_centimos,
                    balance_centimos=balance_centimos,
                    statement_number=statement_number,
                    movement_type=movement_type,
                    movement_status=movement_status,
                    row_hash=row_hash,
                    warnings=[],
                )
            )

        except Exception:
            quarantine_rows += 1

    dates = [r.operation_date for r in parsed_rows if r.operation_date]

    return CajaRuralBankDiagnosticReport(
        source_file=str(source_file_path),
        detected_format="CAJA_RURAL_FLEXIBLE",
        file_sha256=_file_sha256_flexible(source_file_path),
        total_rows=max(len(raw_rows) - header_idx - 1, 0),
        valid_rows=len(parsed_rows),
        quarantine_rows=quarantine_rows,
        income_rows=income_rows,
        expense_rows=expense_rows,
        first_operation_date=min(dates) if dates else None,
        last_operation_date=max(dates) if dates else None,
        total_income_centimos=total_income_centimos,
        total_expense_centimos=total_expense_centimos,
        net_amount_centimos=total_income_centimos + total_expense_centimos,
        by_type=by_type,
        by_status=by_status,
        rows=parsed_rows,
    )



def _insert_or_get_bank_batch(
    conn: sqlite3.Connection,
    *,
    source_file_path: Path,
    report: SantanderBankDiagnosticReport | CajaRuralBankDiagnosticReport | IngBankDiagnosticReport,
    source_type: str,
    bank_label: str,
) -> tuple[int, bool]:
    existing = conn.execute(
        """
        SELECT id
        FROM economic_import_batches
        WHERE source_type = ?
          AND file_sha256 = ?
        LIMIT 1
        """,
        (source_type, report.file_sha256),
    ).fetchone()

    if existing:
        return int(existing["id"]), False

    conn.execute(
        """
        INSERT INTO economic_import_batches (
            source_type,
            source_file_name,
            source_file_path,
            file_sha256,
            detected_format,
            total_rows,
            valid_rows,
            quarantine_rows,
            candidate_payment_rows,
            total_candidate_requested_centimos,
            total_candidate_inserted_centimos,
            total_candidate_dispensed_centimos,
            total_candidate_net_centimos,
            first_start_time,
            last_start_time,
            status,
            notes
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            source_type,
            report.source_file,
            str(source_file_path),
            report.file_sha256,
            report.detected_format,
            report.total_rows,
            report.valid_rows,
            report.quarantine_rows,
            report.income_rows,
            0,
            0,
            0,
            report.net_amount_centimos,
            report.first_operation_date,
            report.last_operation_date,
            "IMPORTED",
            (
                f"Banco {bank_label} importado como movimiento bancario bruto. "
                "No crea cobros, facturas ni vínculos automáticos."
            ),
        ),
    )
    return int(conn.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]), True


def _insert_bank_movement(
    conn: sqlite3.Connection,
    *,
    batch_id: int,
    row: SantanderBankDiagnosticRow | CajaRuralBankDiagnosticRow | IngBankDiagnosticRow,
    bank_name: str,
) -> bool:
    before = conn.total_changes

    statement_number = getattr(row, "statement_number", "")
    account_label = f"apunte:{statement_number}" if statement_number else None

    conn.execute(
        """
        -- Regla crítica:
        -- INSERT OR IGNORE evita reescribir movimientos existentes.
        -- Si el movimiento ya estaba conciliado, se preservan todos los campos linked_*.
        INSERT OR IGNORE INTO bank_movements (
            batch_id,
            row_number,
            row_hash,
            bank_name,
            account_label,
            account_iban,
            operation_date,
            value_date,
            concept,
            amount_centimos,
            balance_centimos,
            currency,
            movement_type,
            movement_status,
            warnings_json,
            linked_client_id,
            linked_expedient_id,
            linked_payment_id,
            linked_by_user_id,
            linked_at,
            link_notes,
            review_status,
            ignored_at,
            ignored_reason
        )
        VALUES (
            ?, ?, ?, ?, ?, ?,
            ?, ?, ?,
            ?, ?, ?,
            ?, ?, ?,
            NULL, NULL, NULL, NULL, NULL, NULL,
            ?, NULL, NULL
        )
        """,
        (
            batch_id,
            row.row_number,
            row.row_hash,
            bank_name,
            account_label,
            None,
            row.operation_date,
            row.value_date,
            row.concept,
            row.amount_centimos,
            row.balance_centimos,
            "EUR",
            row.movement_type,
            row.movement_status,
            __import__("json").dumps(row.warnings, ensure_ascii=False),
            "PENDING_MANUAL_REVIEW",
        ),
    )

    return conn.total_changes > before


def import_santander_bank_file(
    file_path: str | Path,
    *,
    db_path: str | Path = DEFAULT_DB_PATH,
) -> BankImportResult:
    source_file_path = Path(file_path)
    report = diagnose_santander_bank_file(source_file_path)

    inserted = 0
    duplicates = 0

    with connect(db_path) as conn:
        ensure_bank_schema(conn)

        batch_id, batch_created = _insert_or_get_bank_batch(
            conn,
            source_file_path=source_file_path,
            report=report,
            source_type="BANK_SANTANDER",
            bank_label="Santander",
        )

        for row in report.rows:
            if _insert_bank_movement(conn, batch_id=batch_id, row=row, bank_name="SANTANDER"):
                inserted += 1
            else:
                duplicates += 1

        conn.commit()

    return BankImportResult(
        batch_id=batch_id,
        batch_created=batch_created,
        source_file=report.source_file,
        file_sha256=report.file_sha256,
        total_rows=report.total_rows,
        inserted_rows=inserted,
        duplicate_rows=duplicates,
        income_rows=report.income_rows,
        expense_rows=report.expense_rows,
        quarantine_rows=report.quarantine_rows,
        manual_linking_policy=(
            "El banco se importa como movimiento bruto. "
            "No crea cobros, facturas ni vínculos automáticos."
        ),
    )


def import_caja_rural_bank_file(
    file_path: str | Path,
    *,
    db_path: str | Path = DEFAULT_DB_PATH,
) -> BankImportResult:
    source_file_path = Path(file_path)
    try:
        report = diagnose_caja_rural_bank_file_flexible(source_file_path)
    except Exception:
        report = diagnose_caja_rural_bank_file(source_file_path)

    inserted = 0
    duplicates = 0

    with connect(db_path) as conn:
        ensure_bank_schema(conn)

        batch_id, batch_created = _insert_or_get_bank_batch(
            conn,
            source_file_path=source_file_path,
            report=report,
            source_type="BANK_CAJA_RURAL",
            bank_label="Caja Rural",
        )

        for row in report.rows:
            if _insert_bank_movement(conn, batch_id=batch_id, row=row, bank_name="CAJA_RURAL"):
                inserted += 1
            else:
                duplicates += 1

        conn.commit()

    return BankImportResult(
        batch_id=batch_id,
        batch_created=batch_created,
        source_file=report.source_file,
        file_sha256=report.file_sha256,
        total_rows=report.total_rows,
        inserted_rows=inserted,
        duplicate_rows=duplicates,
        income_rows=report.income_rows,
        expense_rows=report.expense_rows,
        quarantine_rows=report.quarantine_rows,
        manual_linking_policy=(
            "Caja Rural se importa como movimiento bruto. "
            "No crea cobros, facturas ni vínculos automáticos."
        ),
    )


def import_ing_bank_file(
    file_path: str | Path,
    *,
    db_path: str | Path = DEFAULT_DB_PATH,
) -> BankImportResult:
    source_file_path = Path(file_path)
    report = diagnose_ing_bank_file(source_file_path)

    inserted = 0
    duplicates = 0

    with connect(db_path) as conn:
        ensure_bank_schema(conn)

        batch_id, batch_created = _insert_or_get_bank_batch(
            conn,
            source_file_path=source_file_path,
            report=report,
            source_type="BANK_ING",
            bank_label="ING",
        )

        for row in report.rows:
            if _insert_bank_movement(conn, batch_id=batch_id, row=row, bank_name="ING"):
                inserted += 1
            else:
                duplicates += 1

        conn.commit()

    return BankImportResult(
        batch_id=batch_id,
        batch_created=batch_created,
        source_file=report.source_file,
        file_sha256=report.file_sha256,
        total_rows=report.total_rows,
        inserted_rows=inserted,
        duplicate_rows=duplicates,
        income_rows=report.income_rows,
        expense_rows=report.expense_rows,
        quarantine_rows=report.quarantine_rows,
        manual_linking_policy=(
            "ING se importa como movimiento bruto. "
            "No crea cobros, facturas ni vínculos automáticos."
        ),
    )


def get_bank_import_summary(
    db_path: str | Path = DEFAULT_DB_PATH,
) -> dict[str, Any]:
    with connect(db_path) as conn:
        ensure_bank_schema(conn)

        batches = conn.execute(
            """
            SELECT
                COUNT(*) AS total_batches,
                COALESCE(SUM(total_rows), 0) AS total_rows,
                COALESCE(SUM(candidate_payment_rows), 0) AS income_rows,
                COALESCE(SUM(quarantine_rows), 0) AS quarantine_rows
            FROM economic_import_batches
            WHERE source_type IN ('BANK_SANTANDER', 'BANK_CAJA_RURAL', 'BANK_ING')
            """
        ).fetchone()

        movements_by_status = conn.execute(
            """
            SELECT movement_status, COUNT(*) AS total
            FROM bank_movements
            GROUP BY movement_status
            ORDER BY movement_status
            """
        ).fetchall()

        movements_by_type = conn.execute(
            """
            SELECT movement_type, COUNT(*) AS total
            FROM bank_movements
            GROUP BY movement_type
            ORDER BY movement_type
            """
        ).fetchall()

        totals = conn.execute(
            """
            SELECT
                COUNT(*) AS total_unique_movements,
                COALESCE(SUM(CASE WHEN amount_centimos > 0 THEN 1 ELSE 0 END), 0) AS income_movements,
                COALESCE(SUM(CASE WHEN amount_centimos < 0 THEN 1 ELSE 0 END), 0) AS expense_movements,
                COALESCE(SUM(CASE WHEN amount_centimos > 0 THEN amount_centimos ELSE 0 END), 0) AS total_income_centimos,
                COALESCE(SUM(CASE WHEN amount_centimos < 0 THEN amount_centimos ELSE 0 END), 0) AS total_expense_centimos,
                COALESCE(SUM(amount_centimos), 0) AS net_amount_centimos,
                COALESCE(SUM(CASE WHEN linked_client_id IS NOT NULL OR linked_expedient_id IS NOT NULL OR linked_payment_id IS NOT NULL THEN 1 ELSE 0 END), 0) AS manual_links_total
            FROM bank_movements
            """
        ).fetchone()

        return {
            "batches": dict(batches or {}),
            "totals": dict(totals or {}),
            "movements_by_status": [dict(row) for row in movements_by_status],
            "movements_by_type": [dict(row) for row in movements_by_type],
        }
